# -*- coding: utf-8 -*-
from lxml import html
import datetime
import requests
import openpyxl

tic = datetime.datetime.now()

workbook = openpyxl.load_workbook('C:/Users/dmore/Downloads/listado2.xlsx')

if 'Todos' not in workbook.sheetnames:
    ws = workbook.create_sheet(title='Todos', index = 0)
    
    headers = ["year","driver","team","car","series","race",
               "article","price","currency","brand","scale","link"]
    ws.append(headers)
# TODO: mejorar el encabezado.
#   - Darle formato de negrita a la primera fila
#   - Darle filtro de datos a la primera fila
#   - Eventualmente darle formato de tabla.
wst = workbook['Todos']


blacklist = []
if 'Blacklist' in workbook.sheetnames:
    ws = workbook['Blacklist']
    for row in ws.rows:
        blacklist.append(row[0].value)

print('\nBlacklist')
print(blacklist)

wishlist = []
if 'Wishlist' in workbook.sheetnames:
    ws = workbook['Wishlist']
    for fila in ws.rows:
        if fila[0].row != 1:
            wishlist.append(fila[0].value)
else:
    workbook.create_sheet(title='Wishlist', index = 1)

wsw = workbook['Wishlist']

print('\nWishlist')
print(wishlist)

if 'Canasta' not in workbook.sheetnames:
    ws = workbook.create_sheet(title='Canasta', index = -1)
    ws.append(['Artículo'])

wsk = workbook['Canasta']
colsWsk = wsk.max_column+1
wsk.cell(1,colsWsk,datetime.datetime.today())

if 'Comentarios' not in workbook.sheetnames:
    workbook.create_sheet(title='Comentarios', index = -1)

wsc = workbook['Comentarios']
wsc.append([datetime.datetime.now()])

page = requests.get("https://ck-modelcars.de/en/l/t-gesamt/k-formel1/scale-1-1-43/a-900/sort-priceup/")
#page = requests.get("https://ck-modelcars.de/en/l/t-gesamt/k-formel1/scale-145-1-24/a-900/sort-priceup/")
page_html = html.fromstring(page.content)

autos_pag = []
for n, car_link in enumerate(page_html.xpath('/html/body/div/div/section/div/div[contains(@class,"div_liste_punkt")]/a/@href')):
    print("{0}: {1}".format(n, car_link))
    car_page = requests.get("https://ck-modelcars.de{}".format(car_link))
    car_html = html.fromstring(car_page.content)
    
    inter_node = car_html.xpath('//h2[contains(@class,"hersteller")]/span')
    if inter_node :
        brand = inter_node[0].text_content()
    else:
        brand = ""
        
    inter_node = car_html.xpath('//h2[contains(@class,"massstab")]/span')
    if inter_node :
        scale = inter_node[0].text_content()
    else:
        scale = ""
   
    inter_node = car_html.xpath('//h6[contains(@class,"team")]/span')
    if inter_node :
        team = inter_node[0].text_content()
    else:
        team = ""
        
    inter_node = car_html.xpath('//h3[contains(@class,"fahrer")]/span')
    if inter_node :
        driver = inter_node[0].text_content()
    else:
        driver = ""
   
    inter_node = car_html.xpath('//h3[contains(@class,"fahrzeug")]/span')
    if inter_node :
        car = inter_node[0].text_content()
    else:
        car = ""
        
    inter_node = car_html.xpath('//h3[contains(@class,"serie")]/span')
    if inter_node :
        series = inter_node[0].text_content()
    else:
        series = ""
      
    inter_node = car_html.xpath('//h6[contains(@class,"saison")]/span')
    if inter_node and inter_node != '':
        year = float(inter_node[0].text_content())
    else:
        year = 0

    inter_node = car_html.xpath('//h6[contains(@class,"full serie")]/span')
    if inter_node :
        race = inter_node[0].text_content()
    else:
        race = ""
        
    inter_node = car_html.xpath('//h2[contains(@class,"artikelnummer")]/span')
    if inter_node :
        article = inter_node[0].text_content()
    else:
        article = ""
        
    inter_node = car_html.xpath('//meta[@itemprop="price"]/@content')
    if inter_node :
        price = float(inter_node[0])
    else:
        price = ""
    
    inter_node = car_html.xpath('//meta[@itemprop="priceCurrency"]/@content')
    if inter_node :
        currency = inter_node[0]
    else:
        currency = ""
    
    autos_pag.append([year,driver,team,car,series,race,article,price,currency,
                  brand,scale,"https://ck-modelcars.de{}".format(car_link)])



autos_nuevos = 0
blacklisted = 0
wishlisted = 0
lista = (fila.value for fila in wst['G'] if fila.row != 1)

for auto in autos_pag:
    ix = -1
    for f in wst['G']:
        if f.value == auto[6]:
           ix = f.row - 1
           break
    
    # si el índice es -1, quiere decir que este es un auto nuevo
    if ix == -1:
        
        # Agrega el auto si es que no está en la blacklist
        if auto[6] not in blacklist:
            wst.append(auto)
            wsc.append(['Nuevo auto: {:4.0f}'.format(auto[0]) + ' ' + auto[1] + ' ' + auto[3] + '(' + auto[6] + ')'])
            autos_nuevos += 1
            print('Nuevo auto: {:4.0f}'.format(auto[0]) + ' ' + auto[1] + ' ' + auto[3] + '(' + auto[6] + ')')
        else:
            blacklisted += 1
    else:
        # si el precio baja, indicarlo en la línea de comandos y en la hoja de "Comentarios"
        if float(auto[7]) < wst['H'][ix].value:
            wsc.append(['bajó de precio: {:4.0f}'.format(auto[0]) + ' ' + auto[1] + ' ' + auto[3] + '(' + auto[6] + ')' +
                       '{} -> {}'.format(wst['H'][ix].value, auto[7])])
            print('bajó de precio: {:4.0f}'.format(auto[0]) + ' ' + auto[1] + ' ' + auto[3] + '(' + auto[6] + ')' +
                       '{} -> {}'.format(wst['H'][ix].value, auto[7]))
        
        # si el precio sube, indicarlo en la línea de comandos y en la hoja de "Comentarios"
        elif float(auto[7]) > wst['H'][ix].value:
            wsc.append(['subió de precio: {:4.0f}'.format(auto[0]) + ' ' + auto[1] + ' ' + auto[3] + '(' + auto[6] + ')' +
                       '{} -> {}'.format(wst['H'][ix].value, auto[7])])
            print('subió de precio: {:4.0f}'.format(auto[0]) + ' ' + auto[1] + ' ' + auto[3] + '(' + auto[6] + ')' +
                       '{} -> {}'.format(wst['H'][ix].value, auto[7]))
        
        
        # actualiza el precio en la hoja "Todos", para tener el precio actual (no importa si está blacklisted)
        wst['H'][ix].value = auto[7]
        
        
    # Agregar a la canasta si es que el auto no está en blacklist (si está en blacklist, ni preocuparse)
    if auto[6] not in blacklist: 
        ix = -1
        for f in wsk['A']:
            if f.value == auto[6]:
               ix = f.row
               break
        
        if ix == -1:
            wsk.cell(wsk.max_row + 1, 1, auto[6])
            wsk.cell(wsk.max_row, wsk.max_column, auto[7])
        else:
            wsk.cell(ix, 1, auto[6])
            wsk.cell(ix, wsk.max_column, auto[7])
            # TODO: Agregar hyperlink (aunque parece que lo agrega solo, ya que está en una tabla)
    
    if auto[6] in wishlist:
        wishlisted += 1

print('Se cargaron {} autos nuevos'.format(autos_nuevos))
wsc.append(['Se cargaron {} autos nuevos'.format(autos_nuevos)])

print('Se vieron {} autos de los {} en la wishlist'.format(wishlisted, len(wishlist)))
wsc.append(['Se vieron {} autos de los {} en la wishlist'.format(wishlisted, len(wishlist))])

print('Se vieron {} autos de los {} en la blacklist'.format(blacklisted, len(blacklist)))
wsc.append(['Se vieron {} autos de los {} en la blacklist'.format(blacklisted, len(blacklist))])

tic = datetime.datetime.now()-tic
print('Revisados {} autos en {}'.format(n,tic))
wsc.append(['Revisados {} autos en {}'.format(n,tic)])

workbook.save(filename = 'C:/Users/dmore/Downloads/listado2.xlsx')
workbook.close()
